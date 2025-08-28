#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2入口カメラのうち1本分を処理する最小構成（もう1本は同じ手順でCSVを出し、後でglobal_id統合）
- 基準フレームで内側ポリゴン＆入場ライン＆アンカー点をクリック
- アンカーをKLTで追跡→ホモグラフィH_t→ポリゴン/ライン射影
- YOLOでperson検出＋簡易トラッキング（ByteTrack）
- 内側のみを対象（ソフトフィルタorハードマスク）
- 外→内ライン交差で1回だけカウント
- CSV出力・オーバーレイ動画生成
"""

import os
import sys
import cv2
import json
import time
import math
import queue
import atexit
import argparse
import datetime as dt
import numpy as np
from shapely.geometry import Point, Polygon, LineString
from PIL import Image, ImageDraw, ImageFont
try:
    from deepface import DeepFace
except Exception:
    DeepFace = None
import torch
from torchvision import models
import torchvision.transforms as T
import subprocess
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# 依存: ultralytics
from ultralytics import YOLO

# ========= ユーティリティ =========

def iso_ts(base_ts, frame_idx, fps):
    """ファイル名等からの開始時刻が与えられたと仮定し、相対でISO化（なければ動画時間を使う）"""
    if base_ts is None:
        # 00:00起点
        seconds = frame_idx / max(fps, 1e-6)
        return (dt.datetime(1970,1,1) + dt.timedelta(seconds=seconds)).isoformat()
    else:
        seconds = frame_idx / max(fps, 1e-6)
        return (base_ts + dt.timedelta(seconds=seconds)).isoformat()

def parse_time_from_filename(path):
    """
    例: camA_20250816_1659.mp4 -> 2025-08-16 16:59:00
    規則が違えばここを書き換え
    """
    name = os.path.basename(path)
    import re
    m = re.search(r'(\d{8})_(\d{4})', name)
    if not m:
        return None
    date_str = m.group(1)  # YYYYMMDD
    time_str = m.group(2)  # HHMM
    try:
        base = dt.datetime.strptime(date_str+time_str, "%Y%m%d%H%M")
        # 日本時間(+09:00)表記にしたければawareにしてもよい
        return base
    except:
        return None

def segments_intersect(p1, p2, q1, q2):
    """2線分の交差判定"""
    def orient(a,b,c):
        return (b[0]-a[0])*(c[1]-a[1]) - (b[1]-a[1])*(c[0]-a[0])
    def onseg(a,b,c):
        return min(a[0],b[0])<=c[0]<=max(a[0],b[0]) and min(a[1],b[1])<=c[1]<=max(a[1],b[1])
    o1=orient(p1,p2,q1); o2=orient(p1,p2,q2); o3=orient(q1,q2,p1); o4=orient(q1,q2,p2)
    if o1==0 and onseg(p1,p2,q1): return True
    if o2==0 and onseg(p1,p2,q2): return True
    if o3==0 and onseg(q1,q2,p1): return True
    if o4==0 and onseg(q1,q2,p2): return True
    return (o1>0) != (o2>0) and (o3>0)!=(o4>0)

def clamp(v, lo, hi):
    return max(lo, min(hi, v))

# ========= クリックGUI（ポリゴン・ライン・アンカー収集） =========

class ClickCollector:
    def __init__(self, window_name, base_img):
        self.win = window_name
        self.base = base_img.copy()
        self.tmp = base_img.copy()
        self.points = []
        self.mode = "polygon"  # polygon -> line -> anchors
        self.line_points = []
        self.anchors = []
        self.help = True
        cv2.namedWindow(self.win, cv2.WINDOW_NORMAL)
        cv2.setMouseCallback(self.win, self._on_mouse)

    def _on_mouse(self, event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            if self.mode == "polygon":
                self.points.append((x,y))
            elif self.mode == "line":
                if len(self.line_points) < 2:
                    self.line_points.append((x,y))
            elif self.mode == "anchors":
                self.anchors.append((x,y))
        if event == cv2.EVENT_RBUTTONDOWN:
            # undo
            if self.mode == "polygon" and self.points:
                self.points.pop()
            elif self.mode == "line" and self.line_points:
                self.line_points.pop()
            elif self.mode == "anchors" and self.anchors:
                self.anchors.pop()

    def draw(self):
        self.tmp = self.base.copy()
        # polygon
        for i,p in enumerate(self.points):
            cv2.circle(self.tmp, p, 4, (0,255,0), -1)
            if i>0:
                cv2.line(self.tmp, self.points[i-1], p, (0,255,0), 2)
        if len(self.points)>=3:
            cv2.line(self.tmp, self.points[-1], self.points[0], (0,255,0), 2)
        # line
        for i,p in enumerate(self.line_points):
            cv2.circle(self.tmp, p, 5, (0,200,255), -1)
        if len(self.line_points)==2:
            cv2.line(self.tmp, self.line_points[0], self.line_points[1], (0,200,255), 2)
        # anchors
        for p in self.anchors:
            cv2.circle(self.tmp, p, 4, (255,0,200), -1)

        # text
        if self.mode == "polygon":
            msg = "Polygon (inside zone): L-click add, R-click undo, ENTER to confirm"
        elif self.mode == "line":
            msg = "Entry line (2 points): L-click add, R-click undo, ENTER to confirm"
        else:
            msg = "Anchors (4-8 pts): L-click add, R-click undo, ENTER to confirm"
        cv2.putText(self.tmp, msg, (10,25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (20,20,240), 2, cv2.LINE_AA)
        cv2.imshow(self.win, self.tmp)

    def run(self):
        while True:
            self.draw()
            k = cv2.waitKey(30) & 0xFF
            if k == 13:  # ENTER
                if self.mode == "polygon" and len(self.points)>=3:
                    self.mode = "line"
                elif self.mode == "line" and len(self.line_points)==2:
                    self.mode = "anchors"
                elif self.mode == "anchors" and len(self.anchors)>=4:
                    break
            elif k in (27, ord('q')): # ESC/Q
                sys.exit(0)
        cv2.destroyWindow(self.win)
        return np.array(self.points,np.float32), np.array(self.line_points,np.float32), np.array(self.anchors,np.float32)

# ========= KLT + ホモグラフィでポリゴン/ライン射影 =========

class HomographyWarp:
    def __init__(self, ref_img, anchors_ref):
        self.prev_gray = cv2.cvtColor(ref_img, cv2.COLOR_BGR2GRAY)
        self.p0 = anchors_ref.reshape(-1,1,2).astype(np.float32)
        self.lk_params = dict(winSize=(21,21), maxLevel=3,
                              criteria=(cv2.TERM_CRITERIA_EPS|cv2.TERM_CRITERIA_COUNT, 30, 0.01))
        self.ok_count = len(anchors_ref)
        self.ref_anchors = anchors_ref.copy()

    def update(self, frame):
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        p1, st, err = cv2.calcOpticalFlowPyrLK(self.prev_gray, gray, self.p0, None, **self.lk_params)
        good_old = self.p0[st==1].reshape(-1,2)
        good_new = p1[st==1].reshape(-1,2)
        H = None
        if len(good_old)>=4:
            H, mask = cv2.findHomography(good_old, good_new, cv2.RANSAC, 3.0)
            self.ok_count = int(mask.sum()) if mask is not None else len(good_old)
        else:
            self.ok_count = 0
        self.prev_gray = gray
        self.p0 = p1
        return H, good_old, good_new

    def reinit(self, frame, anchors_ref):
        """大きくズレた場合に呼ぶ。再クリックしたアンカーで初期化"""
        self.prev_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        self.p0 = anchors_ref.reshape(-1,1,2).astype(np.float32)
        self.ok_count = len(anchors_ref)
        self.ref_anchors = anchors_ref.copy()

def warp_shape(shape_xy, H):
    """N×2のxyをHで射影"""
    pts = shape_xy.reshape(-1,1,2).astype(np.float32)
    out = cv2.perspectiveTransform(pts, H).reshape(-1,2)
    return out

# ========= 検出・追跡 =========

class DetectorTracker:
    """
    Ultralytics YOLO + 内部トラッカー（ByteTrack）に対応
    """
    def __init__(self, model_path="yolov8n.pt", conf=0.25):
        self.model = YOLO(model_path)
        self.conf = conf

    def iter_track(self, source):
        """
        yield per-frame results:
        boxes: list of dict {xyxy:(x1,y1,x2,y2), conf:float, id:int or None}
        """
        results = self.model.track(source=source, classes=[0], conf=self.conf, stream=True, tracker="bytetrack.yaml", verbose=False)
        for r in results:
            boxes = []
            if r.boxes is not None and len(r.boxes)>0:
                # r.boxes.xyxy, r.boxes.conf, r.boxes.id
                xyxy = r.boxes.xyxy.cpu().numpy()
                conf = r.boxes.conf.cpu().numpy()
                ids = r.boxes.id.cpu().numpy() if r.boxes.id is not None else [None]*len(xyxy)
                for (x1,y1,x2,y2),c,i in zip(xyxy, conf, ids):
                    boxes.append(dict(x1=float(x1),y1=float(y1),x2=float(x2),y2=float(y2),conf=float(c),tid=None if i is None else int(i)))
            yield boxes, r.orig_img  # orig_imgはBGR

# ========= カウンティング =========

class EntryCounter:
    def __init__(self, bucket_minutes=15, line_band_px=12):
        self.counted = {}  # tid -> bool
        self.events = []   # appended dict rows
        self.bucket_minutes = bucket_minutes
        self.line_band_px = line_band_px

    def _centroid(self, b):
        return ((b["x1"]+b["x2"])/2.0, (b["y1"]+b["y2"])/2.0)

    def should_count(self, prev_pt, curr_pt, entry_line):
        # entry_line: 2点 (x,y)
        # 太さ(line_band_px)の帯として扱い、外帯->内帯で交差したらTrue
        # シンプルには線分交差で近似
        return segments_intersect(prev_pt, curr_pt, tuple(entry_line[0]), tuple(entry_line[1]))

    def step(self, frame_idx, fps, base_ts, cam_id, boxes_curr, boxes_prev, poly_t, entry_line_t):
        poly = Polygon(poly_t.tolist())
        # soft filter: 内側以外は捨てる
        in_boxes = []
        for b in boxes_curr:
            cx,cy = self._centroid(b)
            if poly.contains(Point(cx,cy)):
                in_boxes.append(b)

        # prev対応
        prev_by_id = {}
        for b in boxes_prev:
            if b["tid"] is not None:
                prev_by_id[b["tid"]] = b

        # 交差判定
        for b in in_boxes:
            tid = b["tid"]
            if tid is None:
                continue
            if tid in self.counted:
                continue
            # 直前は外側か？
            prev = prev_by_id.get(tid)
            if prev is None:
                continue
            prev_c = self._centroid(prev)
            curr_c = self._centroid(b)
            was_out = not poly.contains(Point(prev_c[0], prev_c[1]))
            now_in  = poly.contains(Point(curr_c[0], curr_c[1]))
            if was_out and now_in and self.should_count(prev_c, curr_c, entry_line_t):
                ts = iso_ts(base_ts, frame_idx, fps)
                self.events.append(dict(
                    timestamp=ts,
                    camera_id=cam_id,
                    track_id_local=tid,
                    global_id=None,   # 2カメラ統合時に付与
                    age=None, gender=None, gender_confidence=None,
                    entered_side="outside_to_inside",
                    confidence=None,
                    source_time="filename",
                    notes=""
                ))
                self.counted[tid] = True

    def save_csvs(self, outdir):
        os.makedirs(outdir, exist_ok=True)
        # detail
        import csv
        detail_path = os.path.join(outdir, "entries_detail.csv")
        with open(detail_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=[
                "timestamp","camera_id","track_id_local","global_id",
                "age","gender","gender_confidence","entered_side",
                "confidence","source_time","notes"
            ])
            w.writeheader()
            for e in self.events:
                w.writerow(e)

        # aggregate
        # バケットに丸め
        from collections import defaultdict
        agg = defaultdict(lambda: dict(unique=set(), male=0,female=0,unknown_gender=0,
                                       age_teen=0,age_20s=0,age_30s=0,age_40s=0,age_50s_plus=0,age_unknown=0))
        for e in self.events:
            t = dt.datetime.fromisoformat(e["timestamp"])
            m = self.bucket_minutes
            bucket_start = t.replace(minute=(t.minute//m)*m, second=0, microsecond=0)
            key = bucket_start.isoformat()
            d = agg[key]
            d["unique"].add(e["track_id_local"])
            g = e["gender"]
            if g=="male": d["male"]+=1
            elif g=="female": d["female"]+=1
            else: d["unknown_gender"]+=1
            age = e["age"]
            if age is None: d["age_unknown"]+=1
            elif age<20: d["age_teen"]+=1
            elif age<30: d["age_20s"]+=1
            elif age<40: d["age_30s"]+=1
            elif age<50: d["age_40s"]+=1
            else: d["age_50s_plus"]+=1

        agg_path = os.path.join(outdir, "counts_15min.csv")
        with open(agg_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["bucket_start","bucket_end","unique_visitors","male","female","unknown_gender",
                        "age_teen","age_20s","age_30s","age_40s","age_50s_plus","age_unknown"])
            for k in sorted(agg.keys()):
                start = dt.datetime.fromisoformat(k)
                end = start + dt.timedelta(minutes=self.bucket_minutes)
                d = agg[k]
                w.writerow([start.isoformat(), end.isoformat(), len(d["unique"]),
                            d["male"], d["female"], d["unknown_gender"],
                            d["age_teen"], d["age_20s"], d["age_30s"], d["age_40s"], d["age_50s_plus"], d["age_unknown"]])

        print(f"[SAVE] {detail_path}")
        print(f"[SAVE] {agg_path}")

# ========= 顔・属性・埋め込み =========

class FaceEmbedder:
    """人の再識別の基礎用に、汎用の画像埋め込み（ResNet50のavgpool）を保存。
    本格ReIDモデルではないが、コサイン類似度ベースの後処理の手がかりにする。"""
    def __init__(self):
        self.device = "cpu"
        self.model = models.resnet50(weights=models.ResNet50_Weights.IMAGENET1K_V2)
        self.model.fc = torch.nn.Identity()
        self.model.eval()
        self.model.to(self.device)
        self.tf = T.Compose([
            T.ToPILImage(),
            T.Resize((224,224), interpolation=T.InterpolationMode.BILINEAR),
            T.ToTensor(),
            T.Normalize(mean=[0.485,0.456,0.406], std=[0.229,0.224,0.225]),
        ])

    @torch.inference_mode()
    def embed(self, bgr_img):
        rgb = cv2.cvtColor(bgr_img, cv2.COLOR_BGR2RGB)
        x = self.tf(rgb).unsqueeze(0).to(self.device)
        feat = self.model(x).cpu().numpy().reshape(-1)
        # L2正規化
        n = np.linalg.norm(feat) + 1e-9
        return (feat / n).astype(np.float32)


class AgeGenderEstimator:
    def __init__(self, age_prototxt=None, age_caffemodel=None, gender_prototxt=None, gender_caffemodel=None):
        self.enabled = False
        self.age_net = None
        self.gender_net = None
        try:
            if all([age_prototxt, age_caffemodel, gender_prototxt, gender_caffemodel]):
                self.age_net = cv2.dnn.readNetFromCaffe(age_prototxt, age_caffemodel)
                self.gender_net = cv2.dnn.readNetFromCaffe(gender_prototxt, gender_caffemodel)
                self.enabled = True
        except Exception as e:
            print(f"[WARN] 年齢/性別モデルの読み込みに失敗: {e}")

    def predict(self, face_bgr):
        if not self.enabled:
            return None, None, None
        try:
            blob = cv2.dnn.blobFromImage(face_bgr, 1.0, (227,227), (78.4263377603, 87.7689143744, 114.895847746), swapRB=False)
            # gender
            self.gender_net.setInput(blob)
            g_pred = self.gender_net.forward().flatten()
            genders = ["female","male"]
            g_idx = int(np.argmax(g_pred))
            gender = genders[g_idx]
            g_conf = float(g_pred[g_idx])
            # age
            self.age_net.setInput(blob)
            a_pred = self.age_net.forward().flatten()
            bin_to_age = {0:1,1:5,2:10,3:18,4:22,5:28,6:40,7:50}
            a_idx = int(np.argmax(a_pred))
            age = bin_to_age.get(a_idx, None)
            return age, gender, g_conf
        except Exception:
            return None, None, None


class FaceBestKeeper:
    def __init__(self, outdir, faces_dirname="faces", sheet_path=None, max_sheet_faces=100, age_gender_estimator=None, embedder=None, deepface_backend=None):
        self.outdir = outdir
        self.faces_dir = os.path.join(outdir, faces_dirname)
        os.makedirs(self.faces_dir, exist_ok=True)
        self.sheet_path = sheet_path
        self.max_sheet_faces = max_sheet_faces
        self.age_gender = age_gender_estimator
        self.embedder = embedder
        self.deepface_backend = deepface_backend
        self.best = {}
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        self.live_index_path = os.path.join(self.outdir, "faces_index_live.csv")
        self.alias = {}
        self.alias_threshold = 0.92
        self.last_df_frame = {}
        self.first_seen_frame = {}
        self.last_seen_frame = {}  # 追跡の最終フレーム
        self.df_warmup_frames = 2
        self.df_reinfer_interval = 60

    def set_df_params(self, warmup_frames: int, reinfer_interval: int):
        self.df_warmup_frames = max(0, int(warmup_frames))
        self.df_reinfer_interval = max(0, int(reinfer_interval))

    def _score(self, face_img):
        h, w = face_img.shape[:2]
        area = w * h
        sharp = float(cv2.Laplacian(cv2.cvtColor(face_img, cv2.COLOR_BGR2GRAY), cv2.CV_64F).var())
        return area * (1.0 + 0.001 * sharp)

    def canonical_tid(self, tid):
        # resolve alias chain
        seen = set()
        cur = tid
        while cur in self.alias and cur not in seen:
            seen.add(cur)
            cur = self.alias[cur]
        return cur

    def _maybe_update_alias(self, new_tid):
        info = self.best.get(new_tid)
        if info is None or info.get("emb") is None:
            return
        emb_new = info["emb"]
        # compare with other tids
        for other_tid, other in self.best.items():
            if other_tid == new_tid:
                continue
            emb2 = other.get("emb")
            if emb2 is None:
                continue
            sim = float(np.dot(emb_new, emb2))
            if sim >= self.alias_threshold:
                # map newer to older canonical id
                can = self.canonical_tid(other_tid)
                self.alias[new_tid] = can
                break

    def _write_live_face(self, tid):
        info = self.best[tid]
        age_str = "unknown" if info["age"] is None else str(info["age"])
        gender_str = "unknown" if info["gender"] is None else info["gender"]
        # stable filename per tid to allow overwrite
        path = os.path.join(self.faces_dir, f"id{tid}_best.jpg")
        cv2.imwrite(path, info["img"])
        # rewrite live index csv
        import csv
        with open(self.live_index_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["track_id","timestamp","camera_id",
                        "age_cv","gender_cv","gender_conf_cv",
                        "age_deepface","gender_deepface","gender_conf_deepface",
                        "path"])
            for t2, inf2 in sorted(self.best.items(), key=lambda kv: kv[0]):
                p2 = os.path.join(self.faces_dir, f"id{t2}_best.jpg")
                w.writerow([t2, inf2.get("ts"), None, inf2.get("age"), inf2.get("gender"), inf2.get("gender_conf"),
                            inf2.get("age_df"), inf2.get("gender_df"), inf2.get("gender_conf_df"), p2])

    def update_with_frame(self, frame_bgr, boxes, ts_iso, frame_idx=None):
        for b in boxes:
            tid = b.get("tid")
            if tid is None:
                continue
            if tid not in self.first_seen_frame:
                self.first_seen_frame[tid] = frame_idx if frame_idx is not None else 0
            if frame_idx is not None:
                self.last_seen_frame[tid] = frame_idx
            x1, y1, x2, y2 = map(int, [b["x1"], b["y1"], b["x2"], b["y2"]])
            x1 = max(0, x1); y1 = max(0, y1); x2 = max(x1+1, x2); y2 = max(y1+1, y2)
            roi = frame_bgr[y1:y2, x1:x2]
            if roi.size == 0:
                continue
            gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(48,48))
            cand = []
            if len(faces) == 0:
                hh = roi.shape[0]
                fh = max(48, min(hh, (y2 - y1)//4))
                cand.append(roi[0:fh, :])
            else:
                faces = sorted(faces, key=lambda f: f[2]*f[3], reverse=True)
                for (fx,fy,fw,fh) in faces[:1]:
                    m = 0.08
                    x0 = max(0, int(fx - fw*m))
                    y0 = max(0, int(fy - fh*m))
                    x1_ = min(roi.shape[1], int(fx+fw*(1+m)))
                    y1_ = min(roi.shape[0], int(fy+fh*(1+m)))
                    cand.append(roi[y0:y1_, x0:x1_])
            for face in cand:
                if face.size == 0:
                    continue
                sc = self._score(face)
                cur = self.best.get(tid)
                need_df = False
                if frame_idx is None:
                    need_df = True
                else:
                    first = self.first_seen_frame.get(tid, 0)
                    last = self.last_df_frame.get(tid, -10**9)
                    if frame_idx - first < self.df_warmup_frames:
                        need_df = True
                    elif self.df_reinfer_interval == 0:
                        need_df = True
                    elif frame_idx - last >= self.df_reinfer_interval:
                        need_df = True
                age_df = None; gender_df = None; gender_conf_df = None
                if need_df and DeepFace is not None and self.deepface_backend:
                    try:
                        rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
                        res = DeepFace.analyze(rgb, actions=["age","gender"], detector_backend=self.deepface_backend, enforce_detection=False)
                        if isinstance(res, list) and len(res)>0:
                            res = res[0]
                        age_df = int(res.get("age")) if res.get("age") is not None else None
                        dom = res.get("dominant_gender") or res.get("gender")
                        gender_df = dom if isinstance(dom, str) else None
                        gdict = res.get("gender") if isinstance(res.get("gender"), dict) else None
                        if gdict and gender_df in gdict:
                            gender_conf_df = float(gdict[gender_df])
                        self.last_df_frame[tid] = frame_idx if frame_idx is not None else 0
                    except Exception:
                        pass
                emb = None
                if self.embedder is not None:
                    emb = self.embedder.embed(face)
                if cur is None or sc > cur["score"]:
                    self.best[tid] = dict(img=face.copy(), score=sc, ts=ts_iso,
                                          age=None, gender=None, gender_conf=None,
                                          emb=emb, age_df=age_df if age_df is not None else (cur.get('age_df') if cur else None),
                                          gender_df=gender_df if gender_df is not None else (cur.get('gender_df') if cur else None),
                                          gender_conf_df=gender_conf_df if gender_conf_df is not None else (cur.get('gender_conf_df') if cur else None))
                    self._maybe_update_alias(tid)
                    try:
                        self._write_live_face(tid)
                    except Exception:
                        pass

    def _entry_exit_iso(self, tid, fps, base_ts):
        def to_iso(fi):
            if fi is None:
                return None
            return iso_ts(base_ts, fi, fps)
        return to_iso(self.first_seen_frame.get(tid)), to_iso(self.last_seen_frame.get(tid))

    def save_all(self, camera_id, fps=None, base_ts=None):
        import csv
        index_path = os.path.join(self.outdir, "faces_index.csv")
        with open(index_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["track_id","timestamp","camera_id",
                        "age_df","gender_df","gender_conf_df",
                        "entry_time","exit_time","path"])
            for tid in sorted(self.best.keys()):
                info = self.best[tid]
                age_str = "unknown" if info.get("age_df") is None else str(info.get("age_df"))
                gender_str = "unknown" if info.get("gender_df") is None else info.get("gender_df")
                ts_safe = (info.get("ts") or "").replace(":","-")
                fname = f"id{tid}_{ts_safe}_df_{gender_str}_{age_str}.jpg"
                path = os.path.join(self.faces_dir, fname)
                cv2.imwrite(path, info["img"])
                ent_iso, ext_iso = (None, None)
                if fps is not None and base_ts is not None:
                    ent_iso, ext_iso = self._entry_exit_iso(tid, fps, base_ts)
                w.writerow([
                    tid, info.get("ts"), camera_id,
                    info.get("age_df"), info.get("gender_df"), info.get("gender_conf_df"),
                    ent_iso, ext_iso, path
                ])
        embs = {str(tid): self.best[tid]["emb"] for tid in self.best if self.best[tid].get("emb") is not None}
        if len(embs) > 0:
            np.savez(os.path.join(self.outdir, "faces_embeddings.npz"), **embs)
        if self.sheet_path:
            items = list(self.best.items())[:self.max_sheet_faces]
            if len(items) > 0:
                thumb_w, thumb_h = 224, 224
                cols = min(max(1, int(1000 // (thumb_w + 8))), max(1, len(items)))
                rows = int(np.ceil(len(items) / cols))
                sheet = Image.new("RGB", (cols*(thumb_w+8)+8, rows*(thumb_h+68)+8), (245,245,245))
                draw = ImageDraw.Draw(sheet)
                try:
                    font = ImageFont.load_default()
                except:
                    font = None
                for idx, (tid, info) in enumerate(items):
                    r = idx // cols
                    c = idx % cols
                    x0 = 8 + c*(thumb_w+8)
                    y0 = 8 + r*(thumb_h+68)
                    im = cv2.cvtColor(info["img"], cv2.COLOR_BGR2RGB)
                    im = Image.fromarray(im).resize((thumb_w, thumb_h), Image.BILINEAR)
                    sheet.paste(im, (x0, y0))
                    ent_iso, ext_iso = (None, None)
                    if fps is not None and base_ts is not None:
                        ent_iso, ext_iso = self._entry_exit_iso(tid, fps, base_ts)
                    label = f"ID:{tid}  DF:{(info.get('gender_df') or '?')}/{(info.get('age_df') or '?')}\nentry:{ent_iso or '-'}  exit:{ext_iso or '-'}"
                    draw.text((x0, y0+thumb_h+6), label, fill=(20,20,20), font=font)
                sheet.save(self.sheet_path)
                print(f"[SAVE] faces sheet -> {self.sheet_path}")
        # Excelレポート
        xlsx_path = os.path.join(self.outdir, "faces_report.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "faces"
        ws.append(["track_id","image","image_path","timestamp","age_df","gender_df","entry_time","exit_time"])
        # 列幅と行高を画像に合わせる
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 24
        ws.column_dimensions['H'].width = 24
        row = 2
        thumb_w, thumb_h = 160, 160
        tmp_paths = []
        for tid in sorted(self.best.keys()):
            info = self.best[tid]
            ent_iso, ext_iso = (None, None)
            if fps is not None and base_ts is not None:
                ent_iso, ext_iso = self._entry_exit_iso(tid, fps, base_ts)
            # 路径は faces_index.csv と同等のファイル名に合わせる
            age_str = "unknown" if info.get("age_df") is None else str(info.get("age_df"))
            gender_str = "unknown" if info.get("gender_df") is None else info.get("gender_df")
            ts_safe = (info.get("ts") or "").replace(":","-")
            fname = f"id{tid}_{ts_safe}_df_{gender_str}_{age_str}.jpg"
            img_path = os.path.join(self.faces_dir, fname)
            cv2.imwrite(img_path, info["img"])  # 念のため保存
            ws.cell(row=row, column=1, value=int(tid))
            ws.cell(row=row, column=3, value=img_path)
            ws.cell(row=row, column=4, value=info.get("ts"))
            ws.cell(row=row, column=5, value=info.get("age_df"))
            ws.cell(row=row, column=6, value=info.get("gender_df"))
            ws.cell(row=row, column=7, value=ent_iso)
            ws.cell(row=row, column=8, value=ext_iso)
            # 行高を調整（px→point: *0.75）
            ws.row_dimensions[row].height = thumb_h * 0.75
            # 画像をB列に貼り付け
            tmp_path = os.path.join(self.outdir, f"_thumb_x_{tid}.jpg")
            Image.fromarray(cv2.cvtColor(info["img"], cv2.COLOR_BGR2RGB)).resize((thumb_w, thumb_h), Image.BILINEAR).save(tmp_path, quality=90)
            tmp_paths.append(tmp_path)
            xlimg = XLImage(tmp_path)
            ws.add_image(xlimg, f"B{row}")
            row += 1
        wb.save(xlsx_path)
        # ここで一括クリーンアップ
        for p in tmp_paths:
            try:
                os.remove(p)
            except Exception:
                pass
        print(f"[SAVE] faces report -> {xlsx_path}")

# ========= 滞在時間トラッカー =========

class StayTracker:
    def __init__(self):
        self.active = {}  # tid -> start_frame
        self.total_frames = {}  # tid -> accumulated frames

    def update(self, inside_tids, frame_idx):
        inside_set = set(inside_tids)
        # start new segments
        for tid in inside_set:
            if tid not in self.active:
                self.active[tid] = frame_idx
        # finalize segments that left
        to_finalize = [tid for tid in list(self.active.keys()) if tid not in inside_set]
        for tid in to_finalize:
            start = self.active.pop(tid)
            self.total_frames[tid] = self.total_frames.get(tid, 0) + (frame_idx - start)

    def finalize_all(self, last_frame_idx):
        # close any active segments at end
        for tid, start in list(self.active.items()):
            self.total_frames[tid] = self.total_frames.get(tid, 0) + (last_frame_idx - start)
        self.active.clear()

    def save_csv(self, outdir, fps, base_ts):
        import csv
        path = os.path.join(outdir, "stays.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["track_id","duration_seconds","duration_hhmmss"])  # 区間合算
            for tid in sorted(self.total_frames.keys()):
                seconds = int(self.total_frames[tid] / max(fps, 1e-6))
                hh = seconds // 3600
                mm = (seconds % 3600) // 60
                ss = seconds % 60
                w.writerow([tid, seconds, f"{hh:02d}:{mm:02d}:{ss:02d}"])
        print(f"[SAVE] {path}")

# ========= メイン =========

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--video", required=True, help="入力動画パス")
    ap.add_argument("--outdir", default="output", help="出力ディレクトリ")
    ap.add_argument("--model", default="yolov8n.pt", help="YOLO weights")
    ap.add_argument("--conf", type=float, default=0.3, help="YOLO conf threshold")
    ap.add_argument("--cam-id", default="camA", help="camera id label")
    ap.add_argument("--bucket-minutes", type=int, default=15, help="集計バケット分")
    ap.add_argument("--hard-mask", action="store_true", help="起動時から外側黒塗りON")
    ap.add_argument("--speed", type=float, default=1.0, help="初期再生倍率")
    ap.add_argument("--start-sec", type=float, default=0.0, help="開始オフセット秒（サブクリップ作成）")
    ap.add_argument("--duration-sec", type=float, default=0.0, help="処理する秒数（0で最後まで）")
    ap.add_argument("--clicks-json", default=None, help="ゾーン/ライン/アンカー設定のJSON（指定時はGUI省略）")
    # 顔・属性オプション
    ap.add_argument("--age-prototxt", default=None, help="Age deploy.prototxt のパス")
    ap.add_argument("--age-caffemodel", default=None, help="Age .caffemodel のパス")
    ap.add_argument("--gender-prototxt", default=None, help="Gender deploy.prototxt のパス")
    ap.add_argument("--gender-caffemodel", default=None, help="Gender .caffemodel のパス")
    ap.add_argument("--faces-dirname", default="faces", help="顔切り抜き保存ディレクトリ名(outdir配下)")
    ap.add_argument("--faces-sheet", default=None, help="顔一覧画像の保存先パス(未指定なら outdir/faces_sheet.jpg)")
    ap.add_argument("--faces-sheet-max", type=int, default=100, help="顔一覧に載せる最大人数")
    ap.add_argument("--df-warmup-frames", type=int, default=2, help="DeepFaceを初出現時に連続で走らせるフレーム数")
    ap.add_argument("--df-reinfer-interval", type=int, default=60, help="DeepFaceの再推定間隔フレーム（0で毎フレーム）")
    ap.add_argument("--autosave-sec", type=int, default=0, help="定期自動保存の秒間隔(0で無効)")
    ap.add_argument("--log-detections", default=None, help="検出ログCSVの保存先パス(未指定で無効)")
    args = ap.parse_args()

    # 入力ソース（必要ならサブクリップを作成）
    source_path = args.video
    if (getattr(args, 'start_sec', 0.0) and args.start_sec > 0.0) or (getattr(args, 'duration_sec', 0.0) and args.duration_sec > 0.0):
        try:
            os.makedirs(args.outdir, exist_ok=True)
            temp_path = os.path.join(args.outdir, "_temp_clip.mkv")
            cmd = ["ffmpeg", "-hide_banner", "-loglevel", "error", "-y"]
            if getattr(args, 'start_sec', 0.0) and args.start_sec > 0.0:
                cmd += ["-ss", str(float(args.start_sec))]
            if getattr(args, 'duration_sec', 0.0) and args.duration_sec > 0.0:
                cmd += ["-t", str(float(args.duration_sec))]
            cmd += ["-i", args.video, "-c", "copy", temp_path]
            subprocess.run(cmd, check=True)
            print(f"[INFO] using subclip: {temp_path}")
            source_path = temp_path
        except Exception as e:
            print(f"[WARN] subclip failed: {e}")
            source_path = args.video

    cap = cv2.VideoCapture(source_path)
    if not cap.isOpened():
        print("[ERR] cannot open video:", args.video)
        return
    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    W = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    ok, ref = cap.read()
    if not ok:
        print("[ERR] cannot read first frame")
        return

    # クリック設定のロード/取得
    poly_ref = None
    line_ref = None
    anchors_ref = None
    loaded_from = None
    # 優先1: 明示指定
    if args.clicks_json and os.path.isfile(args.clicks_json):
        try:
            with open(args.clicks_json, "r", encoding="utf-8") as f:
                d = json.load(f)
            poly_ref = np.array(d["poly"], np.float32)
            line_ref = np.array(d["line"], np.float32)
            anchors_ref = np.array(d["anchors"], np.float32)
            loaded_from = args.clicks_json
        except Exception as e:
            print(f"[WARN] clicks-json 読み込み失敗: {e}")
            poly_ref = line_ref = anchors_ref = None
    # 優先2: outdirの既存ファイル
    if poly_ref is None:
        default_clicks = os.path.join(args.outdir, "zone_clicks.json")
        if os.path.isfile(default_clicks):
            try:
                with open(default_clicks, "r", encoding="utf-8") as f:
                    d = json.load(f)
                poly_ref = np.array(d["poly"], np.float32)
                line_ref = np.array(d["line"], np.float32)
                anchors_ref = np.array(d["anchors"], np.float32)
                loaded_from = default_clicks
            except Exception as e:
                print(f"[WARN] 既存 zone_clicks.json 読み込み失敗: {e}")
                poly_ref = line_ref = anchors_ref = None

    if poly_ref is not None and len(poly_ref)>=3 and line_ref is not None and len(line_ref)==2 and anchors_ref is not None and len(anchors_ref)>=4:
        print(f"[INFO] clicks loaded from {loaded_from}")
    else:
        # GUIで収集
        collector = ClickCollector("Define Zone/Line/Anchors", ref)
        poly_ref, line_ref, anchors_ref = collector.run()
        if len(poly_ref)<3 or len(line_ref)!=2 or len(anchors_ref)<4:
            print("[ERR] insufficient clicks")
            return
        # 保存
        with open(os.path.join(args.outdir, "zone_clicks.json"), "w", encoding="utf-8") as f:
            json.dump(dict(poly=poly_ref.tolist(), line=line_ref.tolist(), anchors=anchors_ref.tolist()), f, ensure_ascii=False, indent=2)
            print(f"[SAVE] zone_clicks.json -> {os.path.join(args.outdir, 'zone_clicks.json')}")

    # KLTホモグラフィ
    warp = HomographyWarp(ref, anchors_ref)

    # 検出トラッカー
    dettrk = DetectorTracker(args.model, args.conf)

    # オーバーレイ動画
    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    vw = cv2.VideoWriter(os.path.join(args.outdir, "overlay.mp4"), fourcc, fps, (W,H))

    # カウンタ
    counter = EntryCounter(bucket_minutes=args.bucket_minutes, line_band_px=12)

    # 再生状態
    paused = False
    speed = clamp(args.speed, 0.1, 30.0)  # 倍速
    process_stride = 1  # 2で半分の頻度で重処理、見た目の速度向上
    hard_mask = args.hard_mask
    base_ts = parse_time_from_filename(args.video)
    curr_idx = 0
    prev_boxes = []
    # FPS計測とログ準備
    os.makedirs(args.outdir, exist_ok=True)
    perf_path = os.path.join(args.outdir, "perf_fps.csv")
    perf_f = open(perf_path, "w", newline="", encoding="utf-8")
    import csv as _csv
    perf_w = _csv.writer(perf_f)
    perf_w.writerow(["wall_time_iso","frame_idx","inst_fps","avg_fps","speed","stride"])
    t0 = time.time()
    last_t = t0
    frames_done = 0
    next_fps_log = t0 + 1.0
    # 定期自動保存
    next_autosave = time.time() + args.autosave_sec if getattr(args, 'autosave_sec', 0) and args.autosave_sec > 0 else None
    # 検出ログ
    detlog_f = None
    detlog_w = None
    if getattr(args, 'log_detections', None):
        try:
            detlog_f = open(args.log_detections, "w", newline="", encoding="utf-8")
            detlog_w = _csv.writer(detlog_f)
            detlog_w.writerow(["timestamp","frame_idx","track_id","x1","y1","x2","y2","conf","inside","age_df","gender_df"]) 
            print(f"[INFO] detection log -> {args.log_detections}")
        except Exception as e:
            print(f"[WARN] 検出ログを開けませんでした: {e}")
    # 追加: 顔・属性・埋め込みの準備
    # 引数が未定義のため既定値設定
    if not hasattr(args, 'age_prototxt'):
        args.age_prototxt = None
        args.age_caffemodel = None
        args.gender_prototxt = None
        args.gender_caffemodel = None
        args.faces_dirname = 'faces'
        args.faces_sheet = None
        args.faces_sheet_max = 100
        args.clicks_json = None
        args.autosave_sec = 0
        args.log_detections = None
        args.id_merge_csv = None

    age_est = AgeGenderEstimator(
        age_prototxt=args.age_prototxt,
        age_caffemodel=args.age_caffemodel,
        gender_prototxt=args.gender_prototxt,
        gender_caffemodel=args.gender_caffemodel,
    )
    embedder = FaceEmbedder()
    sheet_path = args.faces_sheet or os.path.join(args.outdir, "faces_sheet.jpg")
    face_keeper = FaceBestKeeper(
        outdir=args.outdir,
        faces_dirname=args.faces_dirname,
        sheet_path=sheet_path,
        max_sheet_faces=args.faces_sheet_max,
        age_gender_estimator=age_est,
        embedder=embedder,
        deepface_backend='retinaface' if DeepFace is not None else None,
    )
    face_keeper.set_df_params(warmup_frames=getattr(args, 'df_warmup_frames', 2), reinfer_interval=getattr(args, 'df_reinfer_interval', 60))
    stay = StayTracker()

    # モデルのtrackは自前VideoCaptureではなくsource=pathで回すのが楽なので、
    # ここでは別ストリームを回しつつフレームも受け取って可視化する
    print("[INFO] Running detector+tracker stream ...")
    stream = dettrk.iter_track(source_path)

    # 表示ウィンドウ
    cv2.namedWindow("Overlay", cv2.WINDOW_NORMAL)

    # メインループ
    for (boxes, frame) in stream:
        curr_idx += 1
        # ストライド処理（見た目速度の改善）
        if (curr_idx % process_stride) != 0:
            key = cv2.waitKey(1) & 0xFF
            if key in (ord('q'), 27):
                break
            continue

        # ホモグラフィ更新
        Hmat, good_old, good_new = warp.update(frame)
        if Hmat is None:
            vis = frame.copy()
            cv2.putText(vis, "H: None (anchors lost) - press R to reinit", (10,30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255),2)
            cv2.imshow("Overlay", vis)
            vw.write(vis)
            paused = True  # 即一時停止
            print("[INFO] Anchors lost -> paused")
            key = cv2.waitKey(0) & 0xFF
            if key in (ord('q'), 27):
                break
            elif key == ord('r') or key == ord('R'):
                rec = ClickCollector("Re-Anchor", frame)
                _, _, anchors_new = rec.run()
                warp.reinit(frame, anchors_new)
            elif key == ord(' '):
                paused = False
            while paused:
                k2 = cv2.waitKey(50) & 0xFF
                if k2 == ord(' '):
                    paused = False
                elif k2 in (ord('q'),27):
                    paused = False
                    break
            continue

        # 射影
        poly_t = warp_shape(poly_ref, Hmat)
        line_t = warp_shape(line_ref, Hmat)

        # ハードマスク
        show = frame.copy()
        if hard_mask:
            mask = np.zeros(show.shape[:2], dtype=np.uint8)
            cv2.fillPoly(mask, [poly_t.astype(np.int32)], 255)
            show = cv2.bitwise_and(show, show, mask=mask)

        # カウント
        counter.step(curr_idx, fps, base_ts, args.cam_id, boxes, prev_boxes, poly_t, line_t)
        # 滞在時間更新
        inside_ids = []
        for b in boxes:
            tid = b.get("tid")
            if tid is None: continue
            cx = (b["x1"]+b["x2"]) / 2.0
            cy = (b["y1"]+b["y2"]) / 2.0
            if Polygon(poly_t.tolist()).contains(Point(cx,cy)):
                inside_ids.append(tid)
        stay.update(inside_ids, curr_idx)
        # 検出ログ出力
        if detlog_w is not None:
            ts_iso = iso_ts(base_ts, curr_idx, fps)
            poly_now = Polygon(poly_t.tolist())
            for b in boxes:
                cx = (b["x1"]+b["x2"]) / 2.0
                cy = (b["y1"]+b["y2"]) / 2.0
                inside = poly_now.contains(Point(cx,cy))
                info = face_keeper.best.get(b.get("tid")) if b.get("tid") is not None else None
                detlog_w.writerow([
                    ts_iso, curr_idx, b.get("tid"),
                    b.get("x1"), b.get("y1"), b.get("x2"), b.get("y2"), b.get("conf"),
                    int(inside),
                    None if info is None else info.get("age_df"),
                    None if info is None else info.get("gender_df"),
                ])

        # 顔更新（ログ出力）
        ts_iso = iso_ts(base_ts, curr_idx, fps)
        face_keeper.update_with_frame(frame, boxes, ts_iso, frame_idx=curr_idx)

        # 可視化
        vis = show.copy()
        cv2.polylines(vis, [poly_t.astype(np.int32)], True, (0,255,0), 2)
        cv2.line(vis, tuple(line_t[0].astype(int)), tuple(line_t[1].astype(int)), (0,200,255), 2)
        for p in good_new:
            cv2.circle(vis, tuple(p.astype(int)), 3, (255,0,200), -1)

        for b in boxes:
            x1,y1,x2,y2 = int(b["x1"]),int(b["y1"]),int(b["x2"]),int(b["y2"])
            tid = b["tid"]
            cv2.rectangle(vis, (x1,y1),(x2,y2),(80,160,255),2)
            cv2.putText(vis, f"ID:{tid}", (x1, y1-6), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (80,160,255),2)
            info = face_keeper.best.get(tid)
            if info is not None and (curr_idx % max(1, int(fps))) == 0:
                print(f"[AGE/GENDER] tid={tid} DF=({info.get('gender_df')},{info.get('age_df')}) @ {ts_iso}")

        txt = f"fps:{fps:.1f}  speed:x{speed:.2f}  anchors:{warp.ok_count}  events:{len(counter.events)}  mask:{'HARD' if hard_mask else 'SOFT'}"
        cv2.putText(vis, txt, (10,30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (40,240,40),2)
        cv2.imshow("Overlay", vis)
        vw.write(vis)

        # FPS更新とログ
        now = time.time()
        dt_wall = max(1e-6, now - last_t)
        inst_fps = 1.0 / dt_wall
        frames_done += 1
        avg_fps = frames_done / max(1e-6, (now - t0))
        if now >= next_fps_log:
            try:
                perf_w.writerow([dt.datetime.fromtimestamp(now).isoformat(), curr_idx, f"{inst_fps:.2f}", f"{avg_fps:.2f}", f"{speed:.2f}", process_stride])
                perf_f.flush()
            except Exception:
                pass
            print(f"[FPS] avg={avg_fps:.2f} inst={inst_fps:.2f} (x{speed:.2f})")
            next_fps_log = now + 1.0
        last_t = now

        # 自動保存
        if next_autosave is not None and now >= next_autosave:
            try:
                counter.save_csvs(args.outdir)
                stay.finalize_all(curr_idx)
                stay.save_csv(args.outdir, fps, base_ts)
                # 顔のライブは都度上書きされているので、一覧・埋め込みも随時保存
                face_keeper.save_all(args.cam_id, fps=fps, base_ts=base_ts)
                print("[AUTO-SAVE] periodic save completed")
            except Exception as e:
                print(f"[WARN] autosave failed: {e}")
            next_autosave = now + args.autosave_sec

        key = cv2.waitKey(int(1000/(fps*min(speed,30.0)))) & 0xFF
        if key in (ord('q'), 27):
            break
        elif key == ord(' '):
            paused = not paused
        elif key == ord('+') or key == ord('='):
            speed = clamp(speed*1.25, 0.1, 30.0)
            process_stride = 1 if speed <= 2.0 else (2 if speed <= 8.0 else 3)
            print(f"[SPEED] -> x{speed:.2f}, stride={process_stride}")
        elif key == ord('-') or key == ord('_'):
            speed = clamp(speed/1.25, 0.1, 30.0)
            process_stride = 1 if speed <= 2.0 else (2 if speed <= 8.0 else 3)
            print(f"[SPEED] -> x{speed:.2f}, stride={process_stride}")
        elif key in (ord('1'), ord('2'), ord('3'), ord('4'), ord('5'), ord('6'), ord('7'), ord('8'), ord('9'), ord('0')):
            fixed = {ord('1'):1.0, ord('2'):2.0, ord('3'):3.0, ord('4'):4.0, ord('5'):5.0,
                     ord('6'):6.0, ord('7'):7.0, ord('8'):8.0, ord('9'):9.0, ord('0'):30.0}[key]
            speed = fixed
            process_stride = 1 if speed <= 2.0 else (2 if speed <= 8.0 else 3)
            print(f"[SPEED] -> x{speed:.2f}, stride={process_stride}")
        elif key == ord('h') or key == ord('H'):
            hard_mask = not hard_mask
        elif key == ord('l') or key == ord('L'):
            counter.line_band_px = {12:4, 4:20, 20:12}[counter.line_band_px]
        elif key == ord('r') or key == ord('R'):
            rec = ClickCollector("Re-Anchor", frame)
            _, _, anchors_new = rec.run()
            warp.reinit(frame, anchors_new)
        elif key == ord('s') or key == ord('S'):
            counter.save_csvs(args.outdir)
            stay.finalize_all(curr_idx)
            stay.save_csv(args.outdir, fps, base_ts)
            face_keeper.save_all(args.cam_id)
            print("[SAVE] manual save done")

        # ポーズ中
        while paused:
            k2 = cv2.waitKey(60) & 0xFF
            if k2 == ord(' '): 
                paused = False
            elif k2 in (ord('q'),27):
                paused = False
                break

        prev_boxes = boxes

    # 後処理
    vw.release()
    cv2.destroyAllWindows()
    counter.save_csvs(args.outdir)
    stay.finalize_all(curr_idx)
    stay.save_csv(args.outdir, fps, base_ts)
    # 顔保存・一覧と埋め込み保存
    face_keeper.save_all(args.cam_id)
    # ログのクローズ
    try:
        perf_f.close()
    except Exception:
        pass
    if detlog_f is not None:
        try:
            detlog_f.close()
            print(f"[SAVE] detection log -> {args.log_detections}")
        except Exception:
            pass
    print(f"[SAVE] overlay video -> {os.path.join(args.outdir,'overlay.mp4')}")
    print("[DONE]")

if __name__ == "__main__":
    main()
